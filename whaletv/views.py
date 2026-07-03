import datetime
import os

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse

from .forms import InhabilitacionForm, TelevisorForm
from .models import Inhabilitacion, PinCodeGenerado, RegistroSync, SyncJob, SyncJobItem, Televisor


def login_view(request):
    """Página de login. Inicia sesión usando correo y contraseña."""
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        user = authenticate(request, username=email, password=password)

        if user is not None:
            login(request, user)
            return redirect('home')

        messages.error(request, 'Correo o contraseña incorrectos.')

    return render(request, 'whaletv/login.html')


@login_required
def home_view(request):
    """Página de inicio, visible solo cuando el usuario ha iniciado sesión."""
    return render(request, 'whaletv/home.html')


def logout_view(request):
    """Cierra la sesión y vuelve al login."""
    logout(request)
    return redirect('login')


# ---------------------------------------------------------------------------
# CRUD de Televisores
# ---------------------------------------------------------------------------

@login_required
def televisor_list(request):
    """Lista todos los televisores, con búsqueda opcional."""
    # Recalcula los estados (vencida/inhabilitado) según la fecha de hoy.
    Televisor.refrescar_todos()

    query = request.GET.get('q', '').strip()
    televisores = Televisor.objects.all()

    if query:
        televisores = televisores.filter(
            Q(mac_address__icontains=query)
            | Q(serial_number__icontains=query)
            | Q(numero_credito__icontains=query)
        )

    return render(request, 'whaletv/televisor_list.html', {
        'televisores': televisores,
        'query': query,
    })


@login_required
def televisor_create(request):
    """Crea un televisor nuevo."""
    if request.method == 'POST':
        form = TelevisorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Televisor creado correctamente.')
            return redirect('televisor_list')
    else:
        form = TelevisorForm()

    return render(request, 'whaletv/televisor_form.html', {
        'form': form,
        'titulo': 'Nuevo televisor',
    })


@login_required
def televisor_update(request, pk):
    """Edita un televisor existente."""
    televisor = get_object_or_404(Televisor, pk=pk)

    if request.method == 'POST':
        form = TelevisorForm(request.POST, instance=televisor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Televisor actualizado correctamente.')
            return redirect('televisor_list')
    else:
        form = TelevisorForm(instance=televisor)

    return render(request, 'whaletv/televisor_form.html', {
        'form': form,
        'titulo': 'Editar televisor',
        'televisor': televisor,
    })


@login_required
def televisor_validar(request, pk):
    """Consulta (dry-run) el estado del televisor en el portal real de WhaleTV."""
    televisor = get_object_or_404(Televisor, pk=pk)
    destino = redirect('registro_sync_tv', mac=televisor.mac_address)

    if request.method != 'POST':
        return destino

    from .portal_sync import sincronizar_televisor

    resultado = sincronizar_televisor(televisor, dry_run=True)

    def estado(b):
        if b is None:
            return '¿?'
        return 'Inhabilitado' if b else 'Habilitado'

    if resultado.ok:
        if resultado.cambiaria:
            messages.warning(
                request,
                f'El televisor {televisor.mac_address} está '
                f'{estado(resultado.remoto_inhabilitado)} de forma remota, pero '
                f'{estado(resultado.local_inhabilitado)} en la app. '
                'Conviene sincronizar para que coincidan.',
            )
        else:
            messages.success(
                request,
                f'El televisor {televisor.mac_address} está '
                f'{estado(resultado.remoto_inhabilitado)} de forma remota, igual que en la app. '
                'No hay nada que sincronizar.',
            )
    else:
        messages.error(
            request,
            f'No se pudo validar el televisor {televisor.mac_address}: {resultado.error}',
        )

    return destino


@login_required
def televisor_sincronizar(request, pk):
    """MODO REAL: aplica el estado local (lock + fecha) en el portal de WhaleTV."""
    televisor = get_object_or_404(Televisor, pk=pk)

    if request.method != 'POST':
        return redirect('televisor_list')

    from .portal_sync import sincronizar_televisor

    resultado = sincronizar_televisor(televisor, dry_run=False)

    if resultado.ok:
        RegistroSync.registrar(
            request.user, televisor,
            aplicado=resultado.aplicado, tipo='individual',
        )

    def estado(b):
        if b is None:
            return '¿?'
        return 'Inhabilitado' if b else 'Habilitado'

    if resultado.ok:
        if resultado.aplicado:
            fecha = televisor.fecha_sincronizar
            messages.success(
                request,
                f'[{televisor.mac_address}] SINCRONIZADO → '
                f'quedó {estado(resultado.remoto_inhabilitado)}'
                + (f' · fecha {fecha:%d/%m/%Y}' if fecha else ''),
            )
        else:
            messages.success(
                request,
                f'[{televisor.mac_address}] Ya estaba igual '
                f'({estado(resultado.remoto_inhabilitado)}), no se cambió nada.',
            )
    else:
        messages.error(
            request,
            f'[{televisor.mac_address}] Error sincronizando: {resultado.error}',
        )

    return redirect('televisor_list')


# ---------------------------------------------------------------------------
# Histórico de facturas de un televisor
# ---------------------------------------------------------------------------

@login_required
def televisor_sincronizar_todos(request):
    """MODO REAL: aplica el estado local de TODOS los televisores en el portal."""
    if request.method != 'POST':
        return redirect('televisor_list')

    from .portal_sync import sincronizar_todos

    Televisor.refrescar_todos()
    televisores = list(Televisor.objects.all())

    if not televisores:
        messages.warning(request, 'No hay televisores para sincronizar.')
        return redirect('televisor_list')

    resultados = sincronizar_todos(televisores, dry_run=False)

    for tv, r in resultados:
        if r.ok:
            RegistroSync.registrar(
                request.user, tv, aplicado=r.aplicado, tipo='masivo',
            )

    aplicados = sum(1 for _, r in resultados if r.ok and r.aplicado)
    sin_cambios = sum(1 for _, r in resultados if r.ok and not r.aplicado)
    fallidos = [(tv, r) for tv, r in resultados if not r.ok]

    if not fallidos:
        messages.success(
            request,
            f'Sincronización masiva lista: {aplicados} aplicados, '
            f'{sin_cambios} sin cambios, {len(resultados)} en total.',
        )
    else:
        detalle = '; '.join(f'{tv.mac_address}: {r.error}' for tv, r in fallidos[:5])
        messages.warning(
            request,
            f'Sincronización masiva: {aplicados} aplicados, {sin_cambios} sin cambios, '
            f'{len(fallidos)} con error → {detalle}',
        )

    return redirect('televisor_list')


def _lanzar_sync_job(request, televisores):
    """Crea y lanza en segundo plano un SyncJob para los televisores dados.

    Reutiliza la infraestructura de sincronización masiva (barra de progreso).
    Devuelve (job, ya_activo):
      - (job, False)  → se creó un job nuevo y está corriendo.
      - (job, True)   → ya había un job activo; se devuelve ese (no se duplica).
      - (None, False) → no había televisores que sincronizar.
    """
    import datetime
    import threading

    from django.utils import timezone

    from .portal_sync import ejecutar_job

    televisores = list(televisores)
    if not televisores:
        return None, False

    # Auto-recupera jobs "colgados": si quedó activo pero sin actividad por
    # varios minutos (server reiniciado / hilo muerto), lo damos por cancelado
    # para no bloquear nuevas sincronizaciones.
    limite = timezone.now() - datetime.timedelta(minutes=10)
    SyncJob.objects.filter(
        estado__in=SyncJob.ACTIVOS, actualizado__lt=limite
    ).update(
        estado='cancelado',
        terminado=timezone.now(),
        error='Cancelado automáticamente (sin actividad).',
    )

    # Si ya hay una realmente en curso, vamos a su progreso en vez de duplicar.
    activo = SyncJob.objects.filter(estado__in=SyncJob.ACTIVOS).first()
    if activo:
        return activo, True

    # 1 navegador por TV, con un tope configurable (WHALETV_PORTAL['MAX_WORKERS']).
    tope = int(settings.WHALETV_PORTAL.get('MAX_WORKERS', 6))
    workers = max(1, min(len(televisores), tope))

    job = SyncJob.objects.create(
        total=len(televisores),
        workers=workers,
        usuario=request.user,
        usuario_email=getattr(request.user, 'email', '') or '',
    )
    SyncJobItem.objects.bulk_create([
        SyncJobItem(job=job, televisor=tv, mac=tv.mac_address) for tv in televisores
    ])

    hilo = threading.Thread(
        target=ejecutar_job, args=(job.pk, workers), daemon=True,
    )
    hilo.start()
    return job, False


@login_required
def sync_iniciar(request):
    """Lanza una sincronización masiva en segundo plano (no bloquea la web)."""
    if request.method != 'POST':
        return redirect('televisor_list')

    import datetime
    import threading

    from django.utils import timezone

    from .portal_sync import ejecutar_job

    # Auto-recupera jobs "colgados": si quedó activo pero sin actividad por
    # varios minutos (server reiniciado / hilo muerto), lo damos por cancelado
    # para no bloquear nuevas sincronizaciones.
    limite = timezone.now() - datetime.timedelta(minutes=10)
    SyncJob.objects.filter(
        estado__in=SyncJob.ACTIVOS, actualizado__lt=limite
    ).update(
        estado='cancelado',
        terminado=timezone.now(),
        error='Cancelado automáticamente (sin actividad).',
    )

    # Si ya hay una realmente en curso, vamos a su progreso en vez de duplicar.
    activo = SyncJob.objects.filter(estado__in=SyncJob.ACTIVOS).first()
    if activo:
        messages.info(
            request,
            'Ya hay una sincronización en curso. Puedes cancelarla aquí si se quedó pegada.',
        )
        return redirect('sync_progreso', pk=activo.pk)

    Televisor.refrescar_todos()
    televisores = list(Televisor.objects.all())
    if not televisores:
        messages.warning(request, 'No hay televisores para sincronizar.')
        return redirect('televisor_list')

    # 1 navegador por TV, con un tope configurable (WHALETV_PORTAL['MAX_WORKERS']).
    tope = int(settings.WHALETV_PORTAL.get('MAX_WORKERS', 6))
    workers = max(1, min(len(televisores), tope))

    job = SyncJob.objects.create(
        total=len(televisores),
        workers=workers,
        usuario=request.user,
        usuario_email=getattr(request.user, 'email', '') or '',
    )
    SyncJobItem.objects.bulk_create([
        SyncJobItem(job=job, televisor=tv, mac=tv.mac_address) for tv in televisores
    ])

    hilo = threading.Thread(
        target=ejecutar_job, args=(job.pk, workers), daemon=True,
    )
    hilo.start()

    return redirect('sync_progreso', pk=job.pk)


@login_required
def validacion_iniciar(request):
    """Lanza una VALIDACIÓN masiva (dry-run): consulta el estado de cada TV en
    el portal y lo compara con el de la app, sin modificar nada."""
    if request.method != 'POST':
        return redirect('televisor_list')

    import datetime
    import threading

    from django.utils import timezone

    from .portal_sync import ejecutar_job

    # Auto-recupera jobs "colgados" (igual que en sync_iniciar).
    limite = timezone.now() - datetime.timedelta(minutes=10)
    SyncJob.objects.filter(
        estado__in=SyncJob.ACTIVOS, actualizado__lt=limite
    ).update(
        estado='cancelado',
        terminado=timezone.now(),
        error='Cancelado automáticamente (sin actividad).',
    )

    activo = SyncJob.objects.filter(estado__in=SyncJob.ACTIVOS).first()
    if activo:
        messages.info(
            request,
            'Ya hay un proceso en curso. Puedes cancelarlo aquí si se quedó pegado.',
        )
        return redirect('sync_progreso', pk=activo.pk)

    Televisor.refrescar_todos()
    televisores = list(Televisor.objects.all())
    if not televisores:
        messages.warning(request, 'No hay televisores para validar.')
        return redirect('televisor_list')

    # 1 navegador por TV, con un tope de 4 (1→1, 2→2, 3→3, 4 o más → 4).
    workers = max(1, min(len(televisores), 4))

    job = SyncJob.objects.create(
        tipo='validacion',
        total=len(televisores),
        workers=workers,
        usuario=request.user,
        usuario_email=getattr(request.user, 'email', '') or '',
    )
    SyncJobItem.objects.bulk_create([
        SyncJobItem(job=job, televisor=tv, mac=tv.mac_address) for tv in televisores
    ])

    hilo = threading.Thread(
        target=ejecutar_job, args=(job.pk, workers),
        kwargs={'dry_run': True}, daemon=True,
    )
    hilo.start()

    return redirect('sync_progreso', pk=job.pk)


@login_required
def sync_cambios(request):
    """Sincroniza SOLO los televisores que cambiaron de estado tras una importación.

    Asigna 1 navegador por televisor, con un tope de 4 (1→1, 2→2, 3→3, 4+→4).
    """
    if request.method != 'POST':
        return redirect('televisor_list')

    import datetime
    import threading

    from django.utils import timezone

    from .portal_sync import ejecutar_job

    pks = request.session.get('sync_cambios_pks') or []
    televisores = list(Televisor.objects.filter(pk__in=pks))
    if not televisores:
        messages.info(request, 'No hay cambios pendientes por sincronizar.')
        return redirect('televisor_list')

    # Auto-recupera jobs "colgados" (igual que en sync_iniciar).
    limite = timezone.now() - datetime.timedelta(minutes=10)
    SyncJob.objects.filter(
        estado__in=SyncJob.ACTIVOS, actualizado__lt=limite
    ).update(
        estado='cancelado',
        terminado=timezone.now(),
        error='Cancelado automáticamente (sin actividad).',
    )

    activo = SyncJob.objects.filter(estado__in=SyncJob.ACTIVOS).first()
    if activo:
        messages.info(
            request,
            'Ya hay una sincronización en curso. Puedes cancelarla aquí si se quedó pegada.',
        )
        return redirect('sync_progreso', pk=activo.pk)

    # 1 navegador por TV, con un tope configurable (WHALETV_PORTAL['MAX_WORKERS']).
    tope = int(settings.WHALETV_PORTAL.get('MAX_WORKERS', 6))
    workers = max(1, min(len(televisores), tope))

    job = SyncJob.objects.create(
        total=len(televisores),
        workers=workers,
        usuario=request.user,
        usuario_email=getattr(request.user, 'email', '') or '',
    )
    SyncJobItem.objects.bulk_create([
        SyncJobItem(job=job, televisor=tv, mac=tv.mac_address) for tv in televisores
    ])

    request.session.pop('sync_cambios_pks', None)

    hilo = threading.Thread(
        target=ejecutar_job, args=(job.pk, workers), daemon=True,
    )
    hilo.start()

    return redirect('sync_progreso', pk=job.pk)


@login_required
def sync_cancelar(request, pk=None):
    """Cancela una sincronización (o todas las que estén activas)."""
    if request.method != 'POST':
        return redirect('televisor_list')

    from django.utils import timezone

    activos = SyncJob.objects.filter(estado__in=SyncJob.ACTIVOS)
    if pk:
        activos = activos.filter(pk=pk)

    n = activos.update(
        estado='cancelado',
        terminado=timezone.now(),
        error='Cancelado por el usuario.',
    )

    if n:
        messages.success(
            request,
            'Sincronización cancelada. Ya puedes lanzar una nueva.'
            if n == 1 else
            f'{n} sincronizaciones canceladas. Ya puedes lanzar una nueva.',
        )
    else:
        messages.info(request, 'No había ninguna sincronización en curso.')

    return redirect('televisor_list')


@login_required
def sync_progreso(request, pk):
    """Página con la barra de progreso de la sincronización (hace polling)."""
    job = get_object_or_404(SyncJob, pk=pk)
    return render(request, 'whaletv/sync_progress.html', {'job': job})


@login_required
def sync_estado(request, pk):
    """Barra de progreso simple para la sincronización de un solo televisor.

    Reusa el mismo job y API de polling que la masiva, pero con una vista
    ligera ('Sincronizando estado'), no el anillo de la sincronización masiva.
    """
    job = get_object_or_404(SyncJob, pk=pk)
    tv_pk = request.GET.get('tv') or ''
    return render(request, 'whaletv/sync_estado.html', {'job': job, 'tv_pk': tv_pk})


@login_required
def sync_progreso_api(request, pk):
    """Devuelve el progreso del job en JSON (para el polling)."""
    job = get_object_or_404(SyncJob, pk=pk)
    items = job.items.all()

    ok = items.filter(estado='ok').count()
    error = items.filter(estado='error').count()
    aplicados = items.filter(estado='ok', aplicado=True).count()
    sin_cambios = items.filter(estado='ok', aplicado=False).count()
    coinciden = items.filter(estado='ok', coincide=True).count()
    no_coinciden = items.filter(estado='ok', coincide=False).count()
    procesados = ok + error
    total = job.total or items.count()
    pct = round(procesados * 100 / total) if total else 100

    errores = list(
        items.filter(estado='error').values('mac', 'mensaje')[:100]
    )

    return JsonResponse({
        'tipo': job.tipo,
        'estado': job.estado,
        'total': total,
        'procesados': procesados,
        'ok': ok,
        'error': error,
        'aplicados': aplicados,
        'sin_cambios': sin_cambios,
        'coinciden': coinciden,
        'no_coinciden': no_coinciden,
        'porcentaje': pct,
        'terminado': job.estado in ('terminado', 'error', 'cancelado'),
        'errores': errores,
    })


def _texto_estado(inhabilitado):
    return 'Inhabilitado' if inhabilitado else 'Habilitado'


def _transicion_estado(final_inhabilitado, aplicado):
    """Devuelve (estado_anterior, estado_final) como texto.

    Si la sincronización 'aplicó' un cambio, el estado anterior era el opuesto
    del final; si no aplicó nada, ya estaba igual (anterior == final).
    """
    final = bool(final_inhabilitado)
    antes = (not final) if aplicado else final
    return _texto_estado(antes), _texto_estado(final)


def _excel_header(ws, headers):
    """Escribe la fila de encabezados con el estilo rosa de la app."""
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.append(headers)
    fill = PatternFill('solid', fgColor='F6186A')
    font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')


def _excel_response(wb, nombre_archivo):
    import io
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


@login_required
def sync_job_export(request, pk):
    """Exporta a Excel el resultado de UN trabajo masivo.

    - Sincronización: lista todos los TVs con su transición de estado.
    - Validación: lista SOLO los TVs cuyo estado NO coincide con el portal.
    """
    from openpyxl import Workbook

    job = get_object_or_404(SyncJob, pk=pk)
    items = job.items.select_related('televisor')

    wb = Workbook()
    ws = wb.active

    if job.tipo == 'validacion':
        ws.title = 'No coinciden'
        _excel_header(ws, ['Dirección MAC', 'Serial Number', 'N° Crédito',
                           'Estado remoto', 'Estado en la app'])
        for item in items.filter(estado='ok', coincide=False):
            tv = item.televisor
            ws.append([
                item.mac,
                tv.serial_number if tv else '—',
                tv.numero_credito if tv else '—',
                _texto_estado(item.remoto_inhabilitado) if item.remoto_inhabilitado is not None else '¿?',
                _texto_estado(item.local_inhabilitado) if item.local_inhabilitado is not None else '¿?',
            ])
        for col, ancho in zip('ABCDE', (20, 20, 22, 20, 20)):
            ws.column_dimensions[col].width = ancho
        nombre_archivo = f'validacion_{job.pk}_{job.creado:%Y%m%d_%H%M}.xlsx'
        return _excel_response(wb, nombre_archivo)

    ws.title = 'Sincronización'
    _excel_header(ws, ['Dirección MAC', 'Serial Number', 'N° Crédito',
                       'Estado anterior', 'Estado final', 'Sincronización', 'Mensaje'])
    for item in items:
        tv = item.televisor
        if item.estado == 'ok':
            # El mensaje OK empieza con "Inhabilitado"/"Habilitado" (estado final).
            final = item.mensaje.strip().lower().startswith('inhabilitado')
            antes, despues = _transicion_estado(final, item.aplicado)
        else:
            antes, despues = '—', '—'
        ws.append([
            item.mac,
            tv.serial_number if tv else '—',
            tv.numero_credito if tv else '—',
            antes,
            despues,
            item.get_estado_display(),
            item.mensaje or '—',
        ])
    for col, ancho in zip('ABCDEFG', (20, 20, 22, 16, 16, 14, 44)):
        ws.column_dimensions[col].width = ancho
    nombre_archivo = f'sincronizacion_{job.pk}_{job.creado:%Y%m%d_%H%M}.xlsx'
    return _excel_response(wb, nombre_archivo)


@login_required
def registro_sync_list(request):
    """Bitácora: quién sincronizó cada televisor (nombre + MAC) y cuándo."""
    query = request.GET.get('q', '').strip()
    registros = RegistroSync.objects.select_related('usuario', 'televisor')

    if query:
        registros = registros.filter(
            Q(mac_address__icontains=query)
            | Q(nombre_persona__icontains=query)
            | Q(usuario_email__icontains=query)
        )

    return render(request, 'whaletv/registro_sync_list.html', {
        'registros': registros[:500],
        'query': query,
        'total': registros.count(),
    })


@login_required
def pincode_list(request):
    """Bitácora de Pin Codes generados (MAC + Passcode + Pin Code)."""
    query = request.GET.get('q', '').strip()
    registros = PinCodeGenerado.objects.select_related('usuario')

    if query:
        registros = registros.filter(
            Q(mac_address__icontains=query)
            | Q(pin_code__icontains=query)
            | Q(passcode__icontains=query)
            | Q(usuario_email__icontains=query)
        )

    return render(request, 'whaletv/pincode_list.html', {
        'registros': registros[:500],
        'query': query,
        'total': registros.count(),
    })


@login_required
def registro_sync_export(request):
    """Exporta a Excel TODAS las sincronizaciones (respeta la búsqueda activa)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    query = request.GET.get('q', '').strip()
    registros = RegistroSync.objects.select_related('usuario', 'televisor')
    if query:
        registros = registros.filter(
            Q(mac_address__icontains=query)
            | Q(nombre_persona__icontains=query)
            | Q(usuario_email__icontains=query)
        )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sincronizaciones'

    headers = ['Fecha', 'Quién sincronizó', 'Persona', 'Dirección MAC',
               'Estado anterior', 'Estado final', 'Resultado', 'Tipo']
    ws.append(headers)
    header_fill = PatternFill('solid', fgColor='F6186A')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for r in registros:
        antes, despues = _transicion_estado(r.inhabilitado, r.aplicado)
        ws.append([
            r.creado.strftime('%d/%m/%Y %H:%M'),
            r.usuario_email or '—',
            r.nombre_persona or '—',
            r.mac_address,
            antes,
            despues,
            'Aplicado' if r.aplicado else 'Sin cambios',
            r.get_tipo_display(),
        ])

    for col, ancho in zip('ABCDEFGH', (18, 26, 22, 20, 16, 16, 14, 12)):
        ws.column_dimensions[col].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="sincronizaciones.xlsx"'
    return response


@login_required
def pincode_export(request):
    """Exporta a Excel TODOS los Pin Codes generados (respeta la búsqueda activa)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    query = request.GET.get('q', '').strip()
    pincodes = PinCodeGenerado.objects.select_related('usuario')
    if query:
        pincodes = pincodes.filter(
            Q(mac_address__icontains=query)
            | Q(pin_code__icontains=query)
            | Q(passcode__icontains=query)
            | Q(usuario_email__icontains=query)
        )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Pin Codes'

    headers = ['Fecha', 'Mac Address', 'Código de acceso', 'Código Pin', 'Quién lo generó']
    ws.append(headers)
    header_fill = PatternFill('solid', fgColor='F6186A')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for p in pincodes:
        ws.append([
            p.creado.strftime('%d/%m/%Y %H:%M'),
            p.mac_address,
            p.passcode,
            p.pin_code,
            p.usuario_email or '—',
        ])

    for col, ancho in zip('ABCDE', (18, 20, 16, 16, 26)):
        ws.column_dimensions[col].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="pincodes.xlsx"'
    return response


@login_required
def registro_sync_tv(request, mac):
    """Detalle de un televisor (tarjeta de info). Las sincronizaciones y los
    pin codes se ven en páginas aparte."""
    televisor = Televisor.objects.filter(mac_address=mac).first()
    return render(request, 'whaletv/registro_sync_tv.html', {
        'mac': mac,
        'televisor': televisor,
        'total': RegistroSync.objects.filter(mac_address=mac).count(),
        'total_pincodes': PinCodeGenerado.objects.filter(mac_address=mac).count(),
    })


@login_required
def registro_sync_tv_records(request, mac):
    """Página con la tabla de sincronizaciones de un televisor."""
    registros = RegistroSync.objects.filter(mac_address=mac).select_related('usuario')
    televisor = Televisor.objects.filter(mac_address=mac).first()
    return render(request, 'whaletv/registro_sync_tv_records.html', {
        'registros': registros,
        'mac': mac,
        'televisor': televisor,
        'total': registros.count(),
    })


@login_required
def registro_sync_tv_pincodes(request, mac):
    """Página con la tabla de Pin Codes generados de un televisor."""
    pincodes = PinCodeGenerado.objects.filter(mac_address=mac).select_related('usuario')
    televisor = Televisor.objects.filter(mac_address=mac).first()
    return render(request, 'whaletv/registro_sync_tv_pincodes.html', {
        'pincodes': pincodes,
        'mac': mac,
        'televisor': televisor,
        'total': pincodes.count(),
    })


@login_required
def habilitar_manual(request, mac):
    """Genera un Pin Code en el portal de Locking System usando el Passcode dado.

    Responde JSON (lo consume el modal de la página vía fetch).
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido.'}, status=405)

    passcode = (request.POST.get('passcode') or '').strip()
    if not passcode:
        return JsonResponse({'ok': False, 'error': 'Debes ingresar el Passcode.'}, status=400)

    from .portal_sync import generar_pin_code

    res, pin = generar_pin_code(mac, passcode)

    if res.ok and pin:
        PinCodeGenerado.objects.create(
            usuario=request.user if getattr(request.user, 'pk', None) else None,
            usuario_email=getattr(request.user, 'email', '') or '',
            mac_address=mac,
            passcode=passcode,
            pin_code=pin,
        )
        return JsonResponse({'ok': True, 'pin': pin, 'mac': mac})

    return JsonResponse({
        'ok': False,
        'error': res.error or 'No se pudo generar el Pin Code.',
        'log': res.log[-6:],
    })


@login_required
def habilitar_sincronizar(request, mac):
    """Sincroniza (modo real) el estado + fecha de la factura del TV con el portal.

    Es lo mismo que el botón 'Sincronizar', pero invocado por AJAX desde el modal
    de Habilitar Manual al pulsar 'Listo'. Responde JSON.
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido.'}, status=405)

    televisor = Televisor.objects.filter(mac_address=mac).first()
    if not televisor:
        return JsonResponse({'ok': False, 'error': 'No existe un televisor con esa MAC.'}, status=404)

    from django.urls import reverse

    from .portal_sync import sincronizar_televisor

    resultado = sincronizar_televisor(televisor, dry_run=False)

    def estado(b):
        if b is None:
            return '¿?'
        return 'Inhabilitado' if b else 'Habilitado'

    if resultado.ok:
        RegistroSync.registrar(
            request.user, televisor, aplicado=resultado.aplicado, tipo='individual',
        )
        if resultado.aplicado:
            fecha = televisor.fecha_sincronizar
            messages.success(
                request,
                f'[{televisor.mac_address}] SINCRONIZADO → '
                f'quedó {estado(resultado.remoto_inhabilitado)}'
                + (f' · fecha {fecha:%d/%m/%Y}' if fecha else ''),
            )
        else:
            messages.success(
                request,
                f'[{televisor.mac_address}] Ya estaba igual '
                f'({estado(resultado.remoto_inhabilitado)}), no se cambió nada.',
            )
        return JsonResponse({'ok': True, 'redirect': reverse('televisor_list')})

    messages.error(
        request,
        f'[{televisor.mac_address}] Error sincronizando: {resultado.error}',
    )
    return JsonResponse({'ok': False, 'error': resultado.error or 'No se pudo sincronizar.'})


@login_required
def registro_sync_tv_export(request, mac):
    """Exporta a Excel las sincronizaciones de un televisor (por MAC)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    registros = RegistroSync.objects.filter(mac_address=mac).select_related('usuario')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sincronizaciones'

    headers = ['Fecha', 'Quién sincronizó', 'Persona', 'Dirección MAC',
               'Estado anterior', 'Estado final', 'Resultado', 'Tipo']
    ws.append(headers)
    header_fill = PatternFill('solid', fgColor='F6186A')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for r in registros:
        antes, despues = _transicion_estado(r.inhabilitado, r.aplicado)
        ws.append([
            r.creado.strftime('%d/%m/%Y %H:%M'),
            r.usuario_email or '—',
            r.nombre_persona or '—',
            r.mac_address,
            antes,
            despues,
            'Aplicado' if r.aplicado else 'Sin cambios',
            r.get_tipo_display(),
        ])

    for col, ancho in zip('ABCDEFGH', (18, 26, 22, 20, 16, 16, 14, 12)):
        ws.column_dimensions[col].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = f"sync_{mac.replace(':', '-')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


@login_required
def registro_sync_tv_pincodes_export(request, mac):
    """Exporta a Excel los Pin Codes generados de un televisor (por MAC)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    pincodes = PinCodeGenerado.objects.filter(mac_address=mac).select_related('usuario')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Pin Codes'

    headers = ['Fecha', 'Mac Address', 'Código de acceso', 'Código Pin', 'Quién lo generó']
    ws.append(headers)
    header_fill = PatternFill('solid', fgColor='F6186A')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for p in pincodes:
        ws.append([
            p.creado.strftime('%d/%m/%Y %H:%M'),
            p.mac_address,
            p.passcode,
            p.pin_code,
            p.usuario_email or '—',
        ])

    for col, ancho in zip('ABCDE', (18, 20, 16, 16, 26)):
        ws.column_dimensions[col].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = f"pincodes_{mac.replace(':', '-')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


@login_required
def televisor_historico(request, pk):
    """Lista el histórico de inhabilitaciones del televisor y permite registrar una nueva."""
    televisor = get_object_or_404(Televisor, pk=pk)

    if request.method == 'POST':
        form = InhabilitacionForm(request.POST)
        if form.is_valid():
            inhabilitacion = form.save(commit=False)
            inhabilitacion.televisor = televisor
            inhabilitacion.mac_address = televisor.mac_address
            inhabilitacion.serial_number = televisor.serial_number
            inhabilitacion.save()
            estado = 'Inhabilitado' if inhabilitacion.estado else 'Habilitado'
            # Al guardar, sincronizamos de una vez el estado con el portal
            # (sin pedir confirmación). El front muestra un modal con la barra.
            job, ya_activo = _lanzar_sync_job(request, [televisor])

            es_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            if es_ajax:
                return JsonResponse({
                    'ok': True,
                    'estado': estado,
                    'job': job.pk if job else None,
                    'ya_activo': bool(ya_activo),
                    'api': reverse('sync_progreso_api', args=[job.pk]) if job else '',
                })

            messages.success(request, f'Estado guardado: {estado}.')
            return redirect('televisor_historico', pk=televisor.pk)
    else:
        form = InhabilitacionForm()

    inhabilitaciones = televisor.inhabilitaciones.all()
    return render(request, 'whaletv/televisor_historico.html', {
        'televisor': televisor,
        'inhabilitaciones': inhabilitaciones,
        'form': form,
    })


@login_required
def inhabilitacion_delete(request, pk):
    """Elimina un registro de inhabilitación (recalcula el estado del TV)."""
    inhabilitacion = get_object_or_404(Inhabilitacion, pk=pk)
    tv_id = inhabilitacion.televisor_id
    if request.method == 'POST':
        inhabilitacion.delete()
        messages.success(request, 'Registro de inhabilitación eliminado.')
    return redirect('televisor_historico', pk=tv_id)


# ---------------------------------------------------------------------------
# Importación masiva de inhabilitaciones (CSV / Excel)
# ---------------------------------------------------------------------------

# Sinónimos aceptados para cada columna (encabezados flexibles).
_COLUMNAS_INHABILITACION = {
    'mac_address': {'mac_address', 'mac', 'mac address'},
    'serial_number': {'serial_number', 'serial', 'serial number', 'sn'},
    'estado': {'estado', 'status', 'lock', 'lock_status', 'bloqueo', 'bloqueado',
               'inhabilitacion', 'inhabilitación', 'inhabilitado'},
}

_COLUMNAS_TV = {
    'mac_address': {'mac_address', 'mac', 'mac address'},
    'serial_number': {'serial_number', 'serial', 'serial number', 'sn'},
    'numero_credito': {'numero_credito', 'numero credito', 'número crédito', 'numero de credito',
                       'credito', 'crédito', 'n_credito', 'num_credito', 'nro_credito'},
}

# Valores aceptados en la columna 'estado' del Excel de inhabilitaciones.
# Se conservan los términos antiguos (bloqueado/desbloqueado) por compatibilidad.
_ESTADO_INHABILITADO = {'inhabilitado', 'inhabilitar', 'inhabilitacion', 'inhabilitación',
                        'bloqueado', 'bloqueo', 'bloquear', 'lock', 'locked', 'si', 'sí',
                        'true', '1', 'x', 'yes', 'y'}
_ESTADO_HABILITADO = {'habilitado', 'habilitar', 'habilitacion', 'habilitación',
                      'desbloqueado', 'desbloqueo', 'desbloquear', 'unlock', 'unlocked',
                      'no', 'false', '0', 'n'}


def _mapear_columnas(headers, columnas):
    """Devuelve {nombre_interno: nombre_real_en_archivo}."""
    mapa = {}
    for real in headers:
        clave = str(real).strip().lower()
        for interno, sinonimos in columnas.items():
            if clave in sinonimos:
                mapa[interno] = real
    return mapa


def _leer_dataframe(archivo):
    """Lee un CSV/Excel subido y devuelve un DataFrame (strings)."""
    import pandas as pd

    nombre = archivo.name.lower()
    if nombre.endswith(('.xlsx', '.xls')):
        return pd.read_excel(archivo, dtype=str)
    return pd.read_csv(archivo, dtype=str, sep=None, engine='python')


def _texto(valor):
    """Convierte un valor de celda a texto limpio ('' si vacío/NaN)."""
    if valor is None:
        return ''
    texto = str(valor).strip()
    return '' if texto.lower() in ('nan', 'nat', 'none') else texto


def _parse_estado_inhabilitacion(valor):
    """Interpreta la celda 'estado' del Excel de inhabilitaciones.

    Devuelve True (inhabilitado), False (habilitado) o None si no se reconoce.
    """
    texto = _texto(valor).lower()
    if texto in _ESTADO_INHABILITADO:
        return True
    if texto in _ESTADO_HABILITADO:
        return False
    return None


def _solo_digitos(valor):
    """Deja solo los dígitos de una celda (para el número de crédito).

    Excel a veces guarda números como '123.0'; quitamos ese sufijo antes de
    filtrar. Devuelve '' si no hay dígitos. Trunca a 60 por seguridad.
    """
    texto = _texto(valor)
    if texto.endswith('.0'):
        texto = texto[:-2]
    digitos = ''.join(c for c in texto if c.isdigit())
    return digitos[:60]


@login_required
def televisor_import(request):
    """Importa televisores/compras masivamente desde Excel/CSV (crea o actualiza)."""
    resultado = None

    if request.method == 'POST' and request.FILES.get('archivo'):
        try:
            df = _leer_dataframe(request.FILES['archivo'])
        except Exception as e:  # noqa: BLE001
            messages.error(request, f'No pude leer el archivo: {e}')
            return redirect('televisor_import')

        mapa = _mapear_columnas(df.columns, _COLUMNAS_TV)
        faltantes = [c for c in ('mac_address', 'serial_number') if c not in mapa]
        if faltantes:
            messages.error(
                request,
                'Faltan columnas obligatorias: ' + ', '.join(faltantes)
                + '. Encabezados encontrados: '
                + ', '.join(str(c) for c in df.columns),
            )
            return redirect('televisor_import')

        creados, actualizados, errores = 0, 0, []

        for i, fila in df.iterrows():
            n = i + 2
            mac = _texto(fila[mapa['mac_address']])
            if not mac:
                errores.append(f'Fila {n}: falta MAC.')
                continue

            serial = _texto(fila[mapa['serial_number']])
            if not serial:
                errores.append(f'Fila {n}: falta el número de serie (serial).')
                continue

            datos = {'serial_number': serial}
            if 'numero_credito' in mapa:
                datos['numero_credito'] = _solo_digitos(fila[mapa['numero_credito']])

            tv = Televisor.objects.filter(mac_address__iexact=mac).first()
            if tv is None:
                Televisor.objects.create(
                    mac_address=mac,
                    serial_number=datos.get('serial_number', ''),
                    numero_credito=datos.get('numero_credito', ''),
                )
                creados += 1
            else:
                for campo, valor in datos.items():
                    setattr(tv, campo, valor)
                tv.save()
                actualizados += 1

        resultado = {
            'creados': creados,
            'actualizados': actualizados,
            'errores': errores,
            'total': creados + actualizados,
        }
        if not errores:
            messages.success(
                request,
                f'Importación lista: {creados} creados, {actualizados} actualizados.',
            )
        else:
            messages.warning(
                request,
                f'Importación con avisos: {creados} creados, {actualizados} actualizados, '
                f'{len(errores)} con error.',
            )

    return render(request, 'whaletv/televisor_import.html', {'resultado': resultado})


@login_required
def televisor_plantilla(request):
    """Descarga una plantilla Excel (.xlsx) para importar televisores/compras."""
    import io

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = 'Televisores'
    ws.append(['mac_address', 'serial_number', 'numero_credito'])
    ws.append(['B4:04:29:7E:3A:AA', 'B4:04:29:7E:3A:AA', '1234567890'])
    ws.append(['B4:04:29:7E:3A:BB', 'B4:04:29:7E:3A:BB', '9876543210'])
    # El número de crédito es texto (puede tener hasta 60 dígitos): forzamos el
    # formato de toda la columna a texto para que Excel no lo redondee.
    for fila in ws['C']:
        fila.number_format = '@'
    for col, ancho in zip('ABC', (22, 22, 24)):
        ws.column_dimensions[col].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_televisores.xlsx"'
    return response


@login_required
def inhabilitacion_import(request):
    """Importa inhabilitaciones desde Excel/CSV (Serial, Mac Address, estado).

    Cada fila fija el estado (inhabilitado/habilitado) del televisor de esa MAC.
    Si el TV no existe se crea. Si el estado cambia, se registra una Inhabilitacion.
    """
    resultado = None
    cambios = []

    if request.method == 'POST' and request.FILES.get('archivo'):
        try:
            df = _leer_dataframe(request.FILES['archivo'])
        except Exception as e:  # noqa: BLE001
            messages.error(request, f'No pude leer el archivo: {e}')
            return redirect('inhabilitacion_import')

        mapa = _mapear_columnas(df.columns, _COLUMNAS_INHABILITACION)
        faltan = [c for c in ('mac_address', 'serial_number', 'estado') if c not in mapa]
        if faltan:
            messages.error(
                request,
                'Faltan columnas en el archivo: ' + ', '.join(faltan)
                + '. Encabezados encontrados: ' + ', '.join(str(c) for c in df.columns),
            )
            return redirect('inhabilitacion_import')

        from django.db import transaction

        errores = []

        # 1) Parsear y validar todas las filas (sin tocar la base de datos).
        #    Dedup por MAC: la última fila de cada MAC manda; conserva el
        #    último serial no vacío.
        orden = []        # MAC (mayúsculas) en orden de primera aparición
        deseado = {}      # MAC_up -> {'mac', 'serial', 'estado'}
        for i, fila in df.iterrows():
            n = i + 2  # +2: fila 1 es encabezado, índice base 0
            mac = _texto(fila[mapa['mac_address']])
            serial = _texto(fila[mapa['serial_number']])
            estado = _parse_estado_inhabilitacion(fila[mapa['estado']])

            if not mac:
                errores.append(f'Fila {n}: falta MAC.')
                continue
            if not serial:
                errores.append(f'Fila {n}: falta el número de serie (serial).')
                continue
            if estado is None:
                errores.append(f'Fila {n}: estado inválido (usa "inhabilitado" o "habilitado").')
                continue

            key = mac.upper()
            if key not in deseado:
                orden.append(key)
            prev = deseado.get(key, {})
            deseado[key] = {
                'mac': mac,
                'serial': serial or prev.get('serial', ''),
                'estado': estado,
            }

        # 2) UNA sola consulta para los televisores que ya existen.
        existentes = {}
        if orden:
            macs = [deseado[k]['mac'] for k in orden]
            for tv in Televisor.objects.filter(mac_address__in=macs):
                existentes[tv.mac_address.upper()] = tv

        cambiados = []
        with transaction.atomic():
            # 3) Crear de golpe los televisores nuevos (1 INSERT múltiple).
            nuevos = [
                Televisor(mac_address=deseado[k]['mac'], serial_number=deseado[k]['serial'])
                for k in orden if k not in existentes
            ]
            if nuevos:
                Televisor.objects.bulk_create(nuevos)
                for tv in nuevos:
                    existentes[tv.mac_address.upper()] = tv
            creados = len(nuevos)
            actualizados = len(orden) - creados

            # 4) Calcular qué cambia y preparar inhabilitaciones + updates en bloque.
            #    inhabilitado ya refleja el estado de la última inhabilitación, así
            #    que comparamos contra él (sin consultar inhabilitaciones por TV).
            inhabilitaciones_nuevas = []
            tvs_update = []
            for k in orden:
                d = deseado[k]
                tv = existentes[k]
                serial_cambia = bool(d['serial']) and tv.serial_number != d['serial']
                if serial_cambia:
                    tv.serial_number = d['serial']
                if d['estado'] != tv.inhabilitado:
                    antes = tv.inhabilitado  # estado del que venía
                    inhabilitaciones_nuevas.append(Inhabilitacion(
                        televisor=tv,
                        mac_address=tv.mac_address,
                        serial_number=d['serial'] or tv.serial_number,
                        estado=d['estado'],
                    ))
                    tv.inhabilitado = d['estado']
                    tvs_update.append(tv)
                    cambiados.append((tv, antes))
                elif serial_cambia:
                    tvs_update.append(tv)

            if inhabilitaciones_nuevas:
                Inhabilitacion.objects.bulk_create(inhabilitaciones_nuevas)
            if tvs_update:
                Televisor.objects.bulk_update(tvs_update, ['inhabilitado', 'serial_number'])

        resultado = {
            'creados': creados,
            'actualizados': actualizados,
            'errores': errores,
            'total': creados + actualizados,
        }

        # Si hubo cambios de estado, sincronizamos de una vez con el portal
        # (sin pedir aprobación) y llevamos al usuario a la barra de progreso.
        if cambiados:
            televisores_cambiados = [tv for tv, _ in cambiados]
            job, ya_activo = _lanzar_sync_job(request, televisores_cambiados)
            n = len(televisores_cambiados)
            resumen = (
                f'Importación lista: {creados} creados, {actualizados} actualizados'
                + (f', {len(errores)} con error' if errores else '')
                + f'. {n} televisor{"es" if n != 1 else ""} con cambio de estado.'
            )
            if job and ya_activo:
                messages.info(
                    request,
                    resumen + ' Ya hay una sincronización en curso; los cambios se '
                    'aplicarán en la próxima.',
                )
                return redirect('sync_progreso', pk=job.pk)
            if job:
                messages.success(request, resumen + ' Sincronizando…')
                return redirect('sync_progreso', pk=job.pk)

        cambios = [{
            'mac': tv.mac_address,
            'serial': tv.serial_number,
            'numero_credito': tv.numero_credito,
            'fecha': tv.fecha_sincronizar,
            'antes': antes,            # estado anterior
            'despues': tv.inhabilitado,  # estado nuevo
        } for tv, antes in cambiados]

    return render(request, 'whaletv/inhabilitacion_import.html', {
        'resultado': resultado,
        'cambios': cambios,
        'n_cambios': len(cambios),
    })


@login_required
def inhabilitacion_plantilla(request):
    """Descarga una plantilla Excel (.xlsx) de ejemplo para importar inhabilitaciones."""
    import io

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inhabilitaciones'
    ws.append(['serial_number', 'mac_address', 'estado'])
    ws.append(['B4:04:29:7E:3A:EE', 'B4:04:29:7E:3A:EE', 'inhabilitado'])
    ws.append(['B4:04:29:7E:3A:FF', 'B4:04:29:7E:3A:FF', 'habilitado'])
    for col, ancho in zip('ABC', (22, 22, 16)):
        ws.column_dimensions[col].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_inhabilitaciones.xlsx"'
    return response
